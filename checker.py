import codecs
import re

# Writing to an excel sheet using Python
import openpyxl

#import file path
import os

#get all files from the current directory
from os import walk

#####################################
## logic ## do not update unless u know how to
#####################################
readingFileError = "error reading file.";

def getFileContent(file_name):
	## filename should be "student_name.html" spaces should be underscore
	## if dont do underscore, immediate failure
	try:
		f = codecs.open(file_name, 'r')
		return f.read()
	except:
		print(readingFileError)
		return 0;


def validateContent(substring,content):
	##remove all next lines and tabs
	cleaned_content = re.sub(r"[\n\t]*", "", content)
	## if element exist give marks
	x = re.search(substring,cleaned_content)
	if x:
		return 1
	else:
		return 0

def validateEachTestCase(file_name,my_testcases):
	score = 0
	markingScheme = []
	content = getFileContent(file_name)
	# if program got error give student 0 score
	if(content == 0):
		return 0
	else:
		for testcase in my_testcases:
			element = testcase[0] 
			validation = testcase[1]
			scoreForThisElement = validateContent(validation,content)
			score += scoreForThisElement
			markingScheme.append([element , scoreForThisElement])

		markingScheme.append(['total_score', score ])
		return markingScheme

def execute(wbkName,testcases):
	

	########################
	#1. create headers first
	########################
	def createHeader():
		try:
			wbk = openpyxl.load_workbook(wbkName)
			first_sheet = wbk.sheetnames[0] #select the first sheet
			ws = wbk[first_sheet]

			counter = 1
			for item in testcases:
				headerElement = item[0]
				ws.cell(row=1, column=counter).value = headerElement
				counter += 1

			wbk.save(wbkName)
			wbk.close

		except:
			print('error creating headers spreadsheet')

	createHeader()

	########################
	#2. write all the student names into the excel
	########################
	def writeStudentNameIntoExcel():
		try:
			student_names = []
			arr = os.listdir()

			for name in arr:
				if('.html' in name):
					student_names.append(name)
			return student_names
		except:
			print("error writing student name into spreadsheet")
			return []

	read_student_names = writeStudentNameIntoExcel()

	########################
	#3. give all the student scores
	########################
	def giveScore():
		student_list = []
		##give score and update the array
		for student in read_student_names:
			result = validateEachTestCase(student,testcases)
			## put the student name in the first index
			result.insert(0,student)
			student_list.append(result)
		##now the student_list will store as follows
		## [student_name, [result_array]]
		return student_list

	student_list = giveScore()

	########################
	#4. write into excel spreadsheet the result
	########################
	def writeResultsIntoSpreadsheet():
		try:
			wbk = openpyxl.load_workbook(wbkName)
			first_sheet = wbk.sheetnames[0] #select the first sheet
			ws = wbk[first_sheet]

			row_index = 1
			col_index = 1
			##rows
			for row in student_list:
				current_row = row
				##cols
				for col in current_row:
					headerElement = col[0]
					marks = col[1]
					#print(col)
					#start from row 2
					if(headerElement == 'student_name'):
						continue
					
					if(type(col) is list):
						ws.cell(row=row_index + 1, column=col_index).value = marks
						#print(marks)
					else:
						student_name = col
						ws.cell(row=row_index + 1, column=col_index).value = student_name
						#print(student_name)
					
					col_index += 1

				col_index = 1 # reset index
				row_index += 1 #increment counter

			wbk.save(wbkName)
			wbk.close

		except:
			print('error writing marks into spreadsheet')

	writeResultsIntoSpreadsheet()


#####################################
## program ## add your text cases below here
#####################################

##1. add your test cases in the list, the program will validate your syntax
## against what is in the program

## pls use regex to check
##["TAG", "REGEX"] <- format
testcases = [
## add student
["student_name","DO_NOT_VALIDATE"],
## html - 12 items
["html","<html\s*.*>\s*.*<\/html>"],
["body","<body\s*.*>\s*.*<\/body>"],
["h1","<h1\s*.*>\s*.*<\/h1>"],
["h2","<h2\s*.*>\s*.*<\/h2>"],
["h3","<h3\s*.*>\s*.*<\/h3>"],
["h4","<h4\s*.*>\s*.*<\/h4>"],
["h5","<h5\s*.*>\s*.*<\/h5>"],
["h6","<h6\s*.*>\s*.*<\/h6>"],
["p","<p\s*.*>\s*.*<\/p>"],
["img","<img.*>"],
["div","<div\s*.*>\s*.*<\/div>"],
["span","<span\s*.*>\s*.*<\/span>"],
## css - 3 items
["style","<style\s*.*>\s*.*<\/style>"],
["font color","color"],
["background color","background-color"],
## javascript - 2 items
["script","<script\s*.*>\s*.*<\/script>"],
["function","function"],
]

wbkName = 'student_list.xlsx'


try:
	execute(wbkName,testcases)
	print("successfully executed")
except:
	print('error')


